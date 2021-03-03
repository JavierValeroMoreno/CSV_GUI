import pandas as pd
from openpyxl import load_workbook

cr = 0

df = pd.read_csv("complete_datas.csv", header = 0)
df_headers = df.columns

df = df.sort_values(by=["image"	])

xls_dest = load_workbook(filename = "DatosTest.xlsx")
pag_act = xls_dest.active

xls_row = pag_act.max_row +1
xls_col = pag_act.max_column +1
print(xls_row)
for row in range(4, xls_row):
    if ((row-4)%9 == 0):
        cr += 1
    else:
        for col in range(3, xls_col):
            print("({}, {})".format(row,col))
            pag_act.cell(row,col).value = df.at[row - cr - 4, df_headers[col-2]]

xls_dest.save(filename ="DatosTestS.xlsx")
